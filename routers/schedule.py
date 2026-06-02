from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from datetime import date
from calendar import monthrange
from typing import Optional
import io
from database import get_db
from models import Shift, Schedule, Employee
from schemas import ShiftCreate, ShiftOut, ScheduleCreate, ScheduleOut
from utils.jwt_helper import require_admin

router = APIRouter(prefix="/admin/schedule", tags=["schedule"])


# ── 班別 CRUD ─────────────────────────────────────────────────────────────────

@router.get("/shifts", response_model=list[ShiftOut])
async def list_shifts(
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Shift).order_by(Shift.start_time))
    return result.scalars().all()


MAX_SHIFTS = 5

@router.post("/shifts", response_model=ShiftOut, status_code=201)
async def create_shift(
    body: ShiftCreate,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    count = (await db.execute(select(Shift))).scalars().all()
    if len(count) >= MAX_SHIFTS:
        raise HTTPException(status_code=400, detail=f"班別數量已達上限（{MAX_SHIFTS} 個）")
    shift = Shift(**body.model_dump())
    db.add(shift)
    await db.commit()
    await db.refresh(shift)
    return shift


@router.patch("/shifts/{shift_id}", response_model=ShiftOut)
async def update_shift(
    shift_id: int,
    body: ShiftCreate,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Shift).where(Shift.id == shift_id))
    shift = result.scalar_one_or_none()
    if not shift:
        raise HTTPException(status_code=404, detail="班別不存在")
    shift.name          = body.name
    shift.start_time    = body.start_time
    shift.end_time      = body.end_time
    shift.color         = body.color
    shift.break_minutes = body.break_minutes
    await db.commit()
    await db.refresh(shift)
    return shift


@router.delete("/shifts/{shift_id}", status_code=204)
async def delete_shift(
    shift_id: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Shift).where(Shift.id == shift_id))
    shift = result.scalar_one_or_none()
    if not shift:
        raise HTTPException(status_code=404, detail="班別不存在")
    await db.execute(delete(Schedule).where(Schedule.shift_id == shift_id))
    await db.delete(shift)
    await db.commit()


# ── 排班 CRUD ─────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[ScheduleOut])
async def list_schedules(
    year: int,
    month: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """查詢指定年月的所有排班"""
    from calendar import monthrange
    first = date(year, month, 1)
    last  = date(year, month, monthrange(year, month)[1])

    result = await db.execute(
        select(Schedule, Employee.display_name, Shift.name, Shift.color, Shift.start_time, Shift.end_time)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Shift,    Schedule.shift_id    == Shift.id)
        .where(Schedule.work_date >= first, Schedule.work_date <= last)
        .order_by(Schedule.work_date, Employee.display_name)
    )
    rows = result.all()
    return [
        ScheduleOut(
            id=s.id, employee_id=s.employee_id, shift_id=s.shift_id,
            work_date=s.work_date, note=s.note, is_overtime=s.is_overtime,
            employee_name=emp_name, shift_name=shift_name,
            shift_color=color, start_time=st, end_time=et,
        )
        for s, emp_name, shift_name, color, st, et in rows
    ]


@router.post("/", response_model=ScheduleOut, status_code=201)
async def create_schedule(
    body: ScheduleCreate,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    # 檢查同一員工同一天是否已有排班
    dup = await db.execute(
        select(Schedule).where(
            Schedule.employee_id == body.employee_id,
            Schedule.work_date   == body.work_date,
        )
    )
    if dup.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="該員工當天已有排班")

    s = Schedule(**body.model_dump())
    db.add(s)
    await db.commit()

    # 重新查詢以取得關聯欄位
    row = await db.execute(
        select(Schedule, Employee.display_name, Shift.name, Shift.color, Shift.start_time, Shift.end_time)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Shift,    Schedule.shift_id    == Shift.id)
        .where(Schedule.id == s.id)
    )
    s, emp_name, shift_name, color, st, et = row.one()
    return ScheduleOut(
        id=s.id, employee_id=s.employee_id, shift_id=s.shift_id,
        work_date=s.work_date, note=s.note, is_overtime=s.is_overtime,
        employee_name=emp_name, shift_name=shift_name,
        shift_color=color, start_time=st, end_time=et,
    )


@router.patch("/{schedule_id}/overtime", response_model=ScheduleOut)
async def toggle_overtime(
    schedule_id: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    s = await db.get(Schedule, schedule_id)
    if not s:
        raise HTTPException(status_code=404, detail="排班不存在")
    s.is_overtime = not s.is_overtime
    await db.commit()
    row = await db.execute(
        select(Schedule, Employee.display_name, Shift.name, Shift.color, Shift.start_time, Shift.end_time)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Shift,    Schedule.shift_id    == Shift.id)
        .where(Schedule.id == s.id)
    )
    s, emp_name, shift_name, color, st, et = row.one()
    return ScheduleOut(
        id=s.id, employee_id=s.employee_id, shift_id=s.shift_id,
        work_date=s.work_date, note=s.note, is_overtime=s.is_overtime,
        employee_name=emp_name, shift_name=shift_name,
        shift_color=color, start_time=st, end_time=et,
    )


@router.get("/export")
async def export_schedule(
    year: int,
    month: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """匯出指定年月排班表為 Excel（黑白版，時間不換行）"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    first = date(year, month, 1)
    last  = date(year, month, monthrange(year, month)[1])
    days_in_month = last.day

    # 查詢排班資料
    result = await db.execute(
        select(Schedule, Employee.display_name, Shift.name, Shift.start_time, Shift.end_time)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Shift,    Schedule.shift_id    == Shift.id)
        .where(Schedule.work_date >= first, Schedule.work_date <= last)
        .order_by(Employee.display_name, Schedule.work_date)
    )
    rows = result.all()

    emp_result = await db.execute(select(Employee).where(Employee.is_active == True).order_by(Employee.display_name))
    employees = emp_result.scalars().all()

    lookup: dict = {}
    for s, emp_name, shift_name, st, et in rows:
        lookup[(s.employee_id, s.work_date.day)] = (shift_name, st, et)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月班表"

    # 樣式
    header_fill  = PatternFill("solid", fgColor="1E293B")
    weekend_fill = PatternFill("solid", fgColor="EEEEEE")
    shift_fill   = PatternFill("solid", fgColor="F0F0F0")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center       = Alignment(horizontal="center", vertical="center")
    med          = Side(style="medium", color="000000")
    thin         = Side(style="thin",   color="AAAAAA")
    outer_border = Border(left=med, right=med, top=med, bottom=med)
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    WEEKDAYS     = ["日","一","二","三","四","五","六"]

    # 第一列：標題
    ws.merge_cells(f"A1:{get_column_letter(days_in_month + 1)}1")
    title_cell = ws["A1"]
    title_cell.value     = f"{year} 年 {month} 月 排班表"
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = header_fill
    title_cell.alignment = center
    title_cell.border    = outer_border
    ws.row_dimensions[1].height = 28

    # 第二列：日期標頭
    ws["A2"].value     = "員工姓名"
    ws["A2"].font      = header_font
    ws["A2"].fill      = header_fill
    ws["A2"].alignment = center
    ws["A2"].border    = cell_border
    ws.column_dimensions["A"].width = 14

    for d in range(1, days_in_month + 1):
        col = get_column_letter(d + 1)
        wd  = date(year, month, d).weekday()
        dow = WEEKDAYS[(wd + 1) % 7]
        is_weekend = dow in ("日", "六")
        cell = ws.cell(row=2, column=d + 1, value=f"{d}({dow})")
        cell.font      = Font(bold=True, color="FF4444" if is_weekend else "FFFFFF", size=9)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border
        ws.column_dimensions[col].width = 14   # 寬到不換行
    ws.row_dimensions[2].height = 20

    # 員工資料列
    for row_idx, emp in enumerate(employees, start=3):
        name_cell = ws.cell(row=row_idx, column=1, value=emp.display_name)
        name_cell.font      = Font(bold=True, size=11)
        name_cell.alignment = center
        name_cell.border    = cell_border
        ws.row_dimensions[row_idx].height = 18

        for d in range(1, days_in_month + 1):
            wd = date(year, month, d).weekday()
            dow = WEEKDAYS[(wd + 1) % 7]
            is_weekend = dow in ("日", "六")
            cell = ws.cell(row=row_idx, column=d + 1)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = cell_border
            if (emp.id, d) in lookup:
                shift_name, st, et = lookup[(emp.id, d)]
                cell.value = f"{shift_name} {st}~{et}"
                cell.font  = Font(bold=True, size=9)
                cell.fill  = shift_fill
            elif is_weekend:
                cell.fill = weekend_fill

    # 凍結標題列
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"schedule_{year}{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf")
async def export_schedule_pdf(
    year: int,
    month: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """匯出指定年月排班表為 PDF（日曆格式，橫向 A4，黑白，每頁最多5員工）"""
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import white, black, HexColor
    import os, math

    # ── 中文字型 ──────────────────────────────────────────────────────────
    font_name = "Helvetica"
    font_bold = "Helvetica-Bold"
    font_paths = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CJK", fp))
                font_name = font_bold = "CJK"
            except Exception:
                pass
            break

    # ── 查詢資料 ──────────────────────────────────────────────────────────
    first = date(year, month, 1)
    last  = date(year, month, monthrange(year, month)[1])

    result = await db.execute(
        select(Schedule, Employee.display_name, Shift.name, Shift.start_time, Shift.end_time)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Shift,    Schedule.shift_id    == Shift.id)
        .where(Schedule.work_date >= first, Schedule.work_date <= last)
        .order_by(Schedule.work_date, Employee.display_name)
    )
    rows = result.all()

    # day -> list of (emp_name, shift_name)
    day_map: dict = {}
    for s, emp_name, shift_name, st, et in rows:
        d = s.work_date.day
        day_map.setdefault(d, []).append((emp_name, shift_name))

    # 查詢員工清單（依姓名排序，分頁用）
    emp_result = await db.execute(
        select(Employee).where(Employee.is_active == True).order_by(Employee.display_name)
    )
    all_employees = emp_result.scalars().all()

    # 查詢班別清單（legend 用）
    shifts_result = await db.execute(select(Shift).order_by(Shift.start_time))
    shifts_list = shifts_result.scalars().all()

    # ── 分頁：每頁最多 5 位員工 ───────────────────────────────────────────
    PER_PAGE   = 5
    emp_pages  = [all_employees[i:i+PER_PAGE] for i in range(0, max(len(all_employees), 1), PER_PAGE)]
    total_pages = len(emp_pages)

    # ── 版面常數（橫向 A4）────────────────────────────────────────────────
    PAGE_W, PAGE_H = landscape(A4)   # 841.9 x 595.3 pt
    MARGIN   = 12 * mm
    WEEKDAYS = ["日", "一", "二", "三", "四", "五", "六"]

    first_dow     = (first.weekday() + 1) % 7
    days_in_month = last.day
    num_weeks     = math.ceil((first_dow + days_in_month) / 7)

    usable_w  = PAGE_W - 2 * MARGIN
    TITLE_H   = 14 * mm
    LEGEND_H  = 9  * mm
    WD_H      = 9  * mm
    cal_h     = PAGE_H - 2 * MARGIN - TITLE_H - LEGEND_H - WD_H
    CELL_W    = usable_w / 7
    CELL_H    = cal_h / num_weeks

    # 每格內可容納的 badge 數
    BADGE_H    = 15
    BADGE_PAD  = 2
    DATE_H     = 15
    MAX_BADGES = int((CELL_H - DATE_H - BADGE_PAD) / (BADGE_H + BADGE_PAD))

    # ── 繪製函式 ──────────────────────────────────────────────────────────
    def draw_page(c, page_employees, page_num):
        """繪製單頁（某批員工的月曆）"""
        emp_names = {e.display_name for e in page_employees}

        cal_top_y = PAGE_H - MARGIN - TITLE_H - LEGEND_H - WD_H

        # ─ 標題 ─
        c.setFillColor(black)
        c.setStrokeColor(black)
        c.setLineWidth(1)
        c.rect(MARGIN, PAGE_H - MARGIN - TITLE_H, usable_w, TITLE_H, fill=0, stroke=1)
        c.setFont(font_bold, 15)
        page_label = f"（第 {page_num}/{total_pages} 頁）" if total_pages > 1 else ""
        c.drawCentredString(
            PAGE_W / 2,
            PAGE_H - MARGIN - TITLE_H + 4.5 * mm,
            f"{year} 年 {month} 月 排班表{page_label}"
        )

        # ─ 班別 Legend ─
        legend_y = PAGE_H - MARGIN - TITLE_H - LEGEND_H
        c.setLineWidth(0.5)
        c.rect(MARGIN, legend_y, usable_w, LEGEND_H, fill=0, stroke=1)
        if shifts_list:
            slot_w = usable_w / len(shifts_list)
            for i, sh in enumerate(shifts_list):
                lx = MARGIN + i * slot_w + 5
                ly = legend_y + LEGEND_H / 2 - 4
                c.setFont(font_bold, 10)
                c.drawString(lx + 3, ly + 2, f"{sh.name}　{sh.start_time}–{sh.end_time}")

        # ─ 星期標頭 ─
        wd_y = PAGE_H - MARGIN - TITLE_H - LEGEND_H - WD_H
        c.setLineWidth(0.5)
        for col, wd in enumerate(WEEKDAYS):
            cx = MARGIN + col * CELL_W
            c.rect(cx, wd_y, CELL_W, WD_H, fill=0, stroke=1)
            c.setFont(font_bold, 11)
            c.setFillColor(black)
            c.drawCentredString(cx + CELL_W / 2, wd_y + 2.5 * mm, wd)

        # ─ 日曆格子 ─
        for week in range(num_weeks):
            for col in range(7):
                cell_idx = week * 7 + col
                day = cell_idx - first_dow + 1
                cx  = MARGIN + col * CELL_W
                cy  = cal_top_y - (week + 1) * CELL_H

                # 週末淡灰底
                is_weekend = col == 0 or col == 6
                if is_weekend:
                    c.setFillColor(HexColor("#EEEEEE"))
                    c.rect(cx, cy, CELL_W, CELL_H, fill=1, stroke=0)

                c.setFillColor(black)
                c.setStrokeColor(black)
                c.setLineWidth(0.5)
                c.rect(cx, cy, CELL_W, CELL_H, fill=0, stroke=1)

                if day < 1 or day > days_in_month:
                    continue

                # 日期數字
                c.setFont(font_bold, 11)
                c.setFillColor(black)
                c.drawString(cx + 4, cy + CELL_H - DATE_H + 2, str(day))

                # 找出本頁員工中有排班的
                entries = [
                    (emp_name, shift_name)
                    for emp_name, shift_name in day_map.get(day, [])
                    if emp_name in emp_names
                ]

                badge_y = cy + CELL_H - DATE_H - BADGE_H - BADGE_PAD
                for emp_name, shift_name in entries[:MAX_BADGES]:
                    if badge_y < cy + BADGE_PAD:
                        break
                    bx = cx + 3
                    bw = CELL_W - 6
                    # 黑框 badge
                    c.setFillColor(black)
                    c.setFont(font_name, 10)
                    label = f"{emp_name}({shift_name})"
                    while c.stringWidth(label, font_name, 10) > bw - 4 and len(label) > 2:
                        label = label[:-1]
                    c.drawString(bx + 2, badge_y + 3, label)
                    badge_y -= (BADGE_H + BADGE_PAD)

        # ─ 頁碼 ─
        c.setFont(font_name, 8)
        c.setFillColor(black)
        c.drawCentredString(PAGE_W / 2, MARGIN / 2, f"- {page_num} -")

    # ── 輸出 ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))

    for page_idx, page_employees in enumerate(emp_pages, start=1):
        draw_page(c, page_employees, page_idx)
        if page_idx < total_pages:
            c.showPage()

    c.save()
    buf.seek(0)

    filename = f"schedule_{year}{month:02d}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf-list")
async def export_schedule_pdf_list(
    year: int,
    month: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """匯出指定年月排班列表為 PDF（橫向 A4，日期為列 × 員工為欄，壓縮至約2頁）"""
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import white, black, HexColor
    import os

    # ── 中文字型 ──────────────────────────────────────────────────────────
    font_name = "Helvetica"
    font_bold = "Helvetica-Bold"
    for fp in ["/System/Library/Fonts/PingFang.ttc", "/System/Library/Fonts/STHeiti Medium.ttc",
               "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"]:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CJK", fp))
                font_name = font_bold = "CJK"
            except Exception:
                pass
            break

    # ── 查詢資料 ──────────────────────────────────────────────────────────
    first = date(year, month, 1)
    last  = date(year, month, monthrange(year, month)[1])
    days_in_month = last.day
    WEEKDAYS = ["日","一","二","三","四","五","六"]

    result = await db.execute(
        select(Schedule, Employee.display_name, Shift.name, Shift.start_time, Shift.end_time)
        .join(Employee, Schedule.employee_id == Employee.id)
        .join(Shift,    Schedule.shift_id    == Shift.id)
        .where(Schedule.work_date >= first, Schedule.work_date <= last)
        .order_by(Employee.display_name, Schedule.work_date)
    )
    rows = result.all()

    # 員工清單（依姓名排序）
    emp_result = await db.execute(
        select(Employee).where(Employee.is_active == True).order_by(Employee.display_name)
    )
    employees = emp_result.scalars().all()

    # lookup: (employee_id, day) -> "班名 start~end"
    lookup: dict = {}
    for s, emp_name, shift_name, st, et in rows:
        lookup[(s.employee_id, s.work_date.day)] = f"{shift_name} {st}~{et}"

    # ── 版面（橫向 A4）────────────────────────────────────────────────────
    PAGE_W, PAGE_H = landscape(A4)
    MARGIN   = 12 * mm
    usable_w = PAGE_W - 2 * MARGIN
    usable_h = PAGE_H - 2 * MARGIN

    TITLE_H  = 12 * mm
    HDR_H    = 10 * mm
    DATE_COL = 20 * mm
    WD_COL   = 8  * mm
    EMP_PER_PAGE = 8   # 每頁最多幾位員工

    # 依員工分組分頁，每頁從第1天開始
    emp_groups = [employees[i:i+EMP_PER_PAGE] for i in range(0, max(len(employees), 1), EMP_PER_PAGE)]
    total_pages = len(emp_groups)
    all_days = list(range(1, days_in_month + 1))

    def get_layout(emp_group):
        n = len(emp_group)
        emp_col_w = (usable_w - DATE_COL - WD_COL) / max(n, 1)
        content_h = usable_h - TITLE_H - HDR_H
        row_h = content_h / days_in_month
        row_h = max(row_h, 5.5 * mm)
        row_h = min(row_h, 10 * mm)
        font_size = max(6.0, min(9.0, row_h / mm * 0.82))
        return emp_col_w, row_h, font_size

    def col_x(i, emp_col_w):
        if i == 0: return MARGIN
        if i == 1: return MARGIN + DATE_COL
        return MARGIN + DATE_COL + WD_COL + (i - 2) * emp_col_w

    def draw_list_page(c, emp_group, page_num):
        emp_col_w, row_h, font_size = get_layout(emp_group)
        table_top = PAGE_H - MARGIN - TITLE_H - HDR_H

        # 標題
        c.setFont(font_bold, 13)
        c.setFillColor(black)
        c.rect(MARGIN, PAGE_H - MARGIN - TITLE_H, usable_w, TITLE_H, fill=0, stroke=1)
        label = f"（第 {page_num}/{total_pages} 頁）" if total_pages > 1 else ""
        c.drawCentredString(PAGE_W/2, PAGE_H - MARGIN - TITLE_H + 3.5*mm,
                            f"{year} 年 {month} 月  排班列表{label}")

        # 員工名標頭列
        c.setFillColor(HexColor("#1E293B"))
        c.rect(MARGIN, table_top, usable_w, HDR_H, fill=1, stroke=0)
        c.setFillColor(white)
        hdr_fs = min(font_size, 9)
        c.setFont(font_bold, hdr_fs)
        c.drawCentredString(MARGIN + DATE_COL/2,          table_top + 3*mm, "日期")
        c.drawCentredString(MARGIN + DATE_COL + WD_COL/2, table_top + 3*mm, "週")
        for ei, emp in enumerate(emp_group):
            cx   = col_x(ei + 2, emp_col_w) + emp_col_w / 2
            name = emp.display_name
            while c.stringWidth(name, font_bold, hdr_fs) > emp_col_w - 4 and len(name) > 1:
                name = name[:-1]
            c.drawCentredString(cx, table_top + 3*mm, name)
        # 標頭縱線
        c.setStrokeColor(white)
        c.setLineWidth(0.4)
        c.line(MARGIN + DATE_COL,          table_top, MARGIN + DATE_COL,          table_top + HDR_H)
        c.line(MARGIN + DATE_COL + WD_COL, table_top, MARGIN + DATE_COL + WD_COL, table_top + HDR_H)
        for ei in range(1, len(emp_group)):
            lx = col_x(ei + 2, emp_col_w)
            c.line(lx, table_top, lx, table_top + HDR_H)

        # 資料列（完整31天，從 5/1 開始）
        for ri, day in enumerate(all_days):
            ry = table_top - (ri + 1) * row_h
            wd = WEEKDAYS[(date(year, month, day).weekday() + 1) % 7]
            is_weekend = wd in ("日", "六")

            if is_weekend:
                c.setFillColor(HexColor("#EEEEEE"))
            elif ri % 2 == 0:
                c.setFillColor(HexColor("#F8F8F8"))
            else:
                c.setFillColor(white)
            c.rect(MARGIN, ry, usable_w, row_h, fill=1, stroke=0)

            c.setFillColor(black)
            c.setFont(font_bold, font_size)
            c.drawCentredString(MARGIN + DATE_COL/2,          ry + row_h*0.32, f"{month}/{day}")
            c.setFont(font_name, font_size)
            c.drawCentredString(MARGIN + DATE_COL + WD_COL/2, ry + row_h*0.32, wd)

            for ei, emp in enumerate(emp_group):
                cx  = col_x(ei + 2, emp_col_w) + emp_col_w / 2
                val = lookup.get((emp.id, day), "")
                if val:
                    c.setFont(font_name, font_size)
                    while c.stringWidth(val, font_name, font_size) > emp_col_w - 4 and len(val) > 1:
                        val = val[:-1]
                    c.drawCentredString(cx, ry + row_h*0.32, val)

            c.setStrokeColor(HexColor("#CCCCCC"))
            c.setLineWidth(0.3)
            c.line(MARGIN, ry, MARGIN + usable_w, ry)

        # 外框 + 標頭底線
        bottom_y = table_top - days_in_month * row_h
        total_h  = HDR_H + days_in_month * row_h
        c.setStrokeColor(black)
        c.setLineWidth(0.8)
        c.rect(MARGIN, bottom_y, usable_w, total_h, fill=0, stroke=1)
        c.line(MARGIN, table_top, MARGIN + usable_w, table_top)

        # 縱線
        c.setStrokeColor(HexColor("#AAAAAA"))
        c.setLineWidth(0.4)
        c.line(MARGIN + DATE_COL,          bottom_y, MARGIN + DATE_COL,          table_top + HDR_H)
        c.line(MARGIN + DATE_COL + WD_COL, bottom_y, MARGIN + DATE_COL + WD_COL, table_top + HDR_H)
        for ei in range(1, len(emp_group)):
            lx = col_x(ei + 2, emp_col_w)
            c.line(lx, bottom_y, lx, table_top + HDR_H)

        c.setFont(font_name, 8)
        c.setFillColor(black)
        c.drawCentredString(PAGE_W/2, MARGIN/2, f"- {page_num} -")

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))
    for page_idx, emp_group in enumerate(emp_groups, start=1):
        draw_list_page(c, emp_group, page_idx)
        if page_idx < total_pages:
            c.showPage()

    c.save()
    buf.seek(0)

    filename = f"schedule_list_{year}{month:02d}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.delete("/{schedule_id}", status_code=204)
async def delete_schedule(
    schedule_id: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Schedule).where(Schedule.id == schedule_id))
    s = result.scalar_one_or_none()
    if not s:
        raise HTTPException(status_code=404, detail="排班不存在")
    await db.delete(s)
    await db.commit()
