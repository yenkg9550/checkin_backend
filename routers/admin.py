from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, delete
from datetime import datetime, date, timezone, timedelta
from typing import Optional
import io, calendar
import openpyxl
from passlib.context import CryptContext
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db
from models import (
    Attendance, Employee, Override, CheckType, SystemSettings, AdminUser, PayrollRecord,
    EmployeeSalaryConfig, ShiftSalaryConfig, PayrollDayOverride,
    Position, Schedule, Shift, LeaveType, PositionLeaveType, EmployeeLeaveType,
    LeaveRequest, LeaveRecord,
)
from schemas import (
    AttendanceWithUser, OverrideRequest, EmployeeOut,
    SystemSettingsOut, SystemSettingsUpdate,
    AdminCreateRequest, AdminPasswordUpdate, AdminUserOut, AdminPermissionsUpdate,
    OverrideRequestOut, OverrideApproveReject,
)
from utils.jwt_helper import require_admin, require_super_admin

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

TZ_TAIPEI = timezone(timedelta(hours=8))

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/report", response_model=list[AttendanceWithUser])
async def daily_report(
    report_date: Optional[date] = None,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """查詢指定日期（預設今天台灣時間）所有員工打卡紀錄"""
    target = report_date or datetime.now(TZ_TAIPEI).date()
    # 前端傳台灣日期，轉換為對應 UTC 範圍（UTC = 台灣時間 - 8h）
    day_start = datetime.combine(target, datetime.min.time()) - timedelta(hours=8)
    day_end   = datetime.combine(target, datetime.max.time()) - timedelta(hours=8)

    result = await db.execute(
        select(Attendance, Employee.display_name, Employee.picture_url)
        .join(Employee, Attendance.employee_id == Employee.id)
        .where(and_(Attendance.checked_at >= day_start, Attendance.checked_at <= day_end))
        .order_by(Attendance.checked_at)
    )
    rows = result.all()

    return [
        AttendanceWithUser(
            **{c.key: getattr(att, c.key) for c in Attendance.__table__.columns},
            display_name=name,
            picture_url=pic,
        )
        for att, name, pic in rows
    ]


@router.get("/attendance/monthly")
async def monthly_attendance(
    year:        int,
    month:       int,
    employee_id: Optional[int] = None,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """回傳指定月份的所有打卡紀錄，依員工 + 日期分組為 (clock_in, clock_out) 配對。"""
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    win_start = datetime.combine(first_day, datetime.min.time()) - timedelta(hours=8)
    win_end   = datetime.combine(last_day,  datetime.max.time()) - timedelta(hours=8)

    q = (
        select(Attendance, Employee.display_name, Employee.picture_url)
        .join(Employee, Attendance.employee_id == Employee.id)
        .where(Attendance.checked_at >= win_start)
        .where(Attendance.checked_at <= win_end)
        .order_by(Employee.display_name, Attendance.checked_at)
    )
    if employee_id:
        q = q.where(Attendance.employee_id == employee_id)

    rows = (await db.execute(q)).all()

    # 撈排班休息時間：(employee_id, work_date) → break_minutes
    from models import Schedule, Shift as ShiftModel
    sched_q = (
        select(Schedule.employee_id, Schedule.work_date, ShiftModel.break_minutes)
        .join(ShiftModel, Schedule.shift_id == ShiftModel.id)
        .where(Schedule.work_date >= first_day, Schedule.work_date <= last_day)
    )
    if employee_id:
        sched_q = sched_q.where(Schedule.employee_id == employee_id)
    sched_rows = (await db.execute(sched_q)).all()
    break_map: dict[tuple, int] = {}
    sched_set: set[tuple] = set()
    for r in sched_rows:
        key = (r.employee_id, r.work_date.isoformat())
        break_map[key] = r.break_minutes
        sched_set.add(key)

    # 依員工 + 台灣日期分組，配對 clock_in / clock_out
    from collections import defaultdict
    grouped: dict[tuple, dict] = defaultdict(lambda: {
        "clock_in": None, "clock_out": None,
        "clock_in_id": None, "clock_out_id": None,
    })
    emp_info: dict[int, dict] = {}

    for att, name, pic in rows:
        eid = att.employee_id
        emp_info[eid] = {"employee_id": eid, "employee_name": name, "picture_url": pic}
        tw_dt = att.checked_at + timedelta(hours=8)
        key = (eid, tw_dt.date().isoformat())
        if att.check_type == CheckType.clock_in:
            grouped[key]["clock_in"]    = (att.checked_at + timedelta(hours=8)).strftime("%Y-%m-%dT%H:%M:%S")
            grouped[key]["clock_in_id"] = att.id
        else:
            grouped[key]["clock_out"]    = (att.checked_at + timedelta(hours=8)).strftime("%Y-%m-%dT%H:%M:%S")
            grouped[key]["clock_out_id"] = att.id

    result = []
    for (eid, work_date), times in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
        info = emp_info[eid]
        ci_str = times["clock_in"]
        co_str = times["clock_out"]
        worked_min = None
        if ci_str and co_str:
            ci = datetime.fromisoformat(ci_str)
            co = datetime.fromisoformat(co_str)
            diff = (co - ci).total_seconds() / 60
            if diff < 0:
                diff += 1440  # overnight
            break_min = break_map.get((eid, work_date), 0)
            worked_min = max(0, int(diff) - break_min)
        result.append({
            **info,
            "work_date":      work_date,
            "clock_in":       ci_str,
            "clock_out":      co_str,
            "clock_in_id":    times["clock_in_id"],
            "clock_out_id":   times["clock_out_id"],
            "worked_minutes": worked_min,
            "has_schedule":   (eid, work_date) in sched_set,
        })

    return result


@router.patch("/attendance/{att_id}")
async def update_attendance(
    att_id: int,
    body: dict,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """更新單筆打卡紀錄的時間（台灣時間 HH:MM 字串）"""
    result = await db.execute(select(Attendance).where(Attendance.id == att_id))
    att = result.scalar_one_or_none()
    if not att:
        raise HTTPException(status_code=404, detail="打卡紀錄不存在")
    # body: { "time": "HH:MM", "date": "YYYY-MM-DD" }
    time_str = body.get("time")  # e.g. "08:30"
    date_str = body.get("date")  # e.g. "2026-05-05"
    if not time_str or not date_str:
        raise HTTPException(status_code=422, detail="需提供 date 和 time")
    h, m = map(int, time_str.split(":"))
    tw_dt = datetime.fromisoformat(date_str).replace(hour=h, minute=m, second=0, microsecond=0)
    att.checked_at = tw_dt - timedelta(hours=8)
    await db.commit()
    return {"success": True}


@router.delete("/attendance/{att_id}")
async def delete_attendance(
    att_id: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """刪除單筆打卡紀錄"""
    result = await db.execute(select(Attendance).where(Attendance.id == att_id))
    att = result.scalar_one_or_none()
    if not att:
        raise HTTPException(status_code=404, detail="打卡紀錄不存在")
    await db.delete(att)
    await db.commit()
    return {"success": True}


@router.get("/employees", response_model=list[EmployeeOut])
async def list_employees(
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Employee).order_by(Employee.created_at))
    return result.scalars().all()


@router.patch("/employees/{employee_id}/role")
async def update_role(
    employee_id: int,
    role: str,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """設定員工角色（employee / admin）"""
    result = await db.execute(select(Employee).where(Employee.id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="員工不存在")
    emp.role = role
    await db.commit()
    return {"success": True}


@router.patch("/employees/{employee_id}/hire-date")
async def update_hire_date(
    employee_id: int,
    hire_date: str,   # YYYY-MM-DD or ""
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """設定員工到職日"""
    from datetime import date as date_cls
    result = await db.execute(select(Employee).where(Employee.id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="員工不存在")
    emp.hire_date = date_cls.fromisoformat(hire_date) if hire_date else None
    await db.commit()
    return {"success": True}


@router.get("/overrides")
async def list_overrides(
    employee_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """查詢補打卡紀錄，可依員工、年月篩選"""
    q = (
        select(Override, Employee.display_name, Employee.picture_url)
        .join(Employee, Override.employee_id == Employee.id)
        .order_by(Override.created_at.desc())
    )
    if employee_id:
        q = q.where(Override.employee_id == employee_id)
    if year and month:
        import calendar as cal
        last_day = cal.monthrange(year, month)[1]
        # override_at 儲存 UTC，月份範圍轉回 UTC
        m_start = datetime(year, month, 1) - timedelta(hours=8)
        m_end   = datetime(year, month, last_day, 23, 59, 59) - timedelta(hours=8)
        q = q.where(Override.override_at >= m_start, Override.override_at <= m_end)
    result = await db.execute(q)
    rows = result.all()
    return [
        {
            "id": ov.id,
            "employee_id": ov.employee_id,
            "display_name": name,
            "picture_url": pic,
            "check_type": ov.check_type,
            "override_at": ov.override_at.replace(tzinfo=timezone.utc).astimezone(TZ_TAIPEI).isoformat(),
            "reason": ov.reason,
            "created_at": ov.created_at.replace(tzinfo=timezone.utc).astimezone(TZ_TAIPEI).isoformat(),
        }
        for ov, name, pic in rows
    ]


@router.post("/override")
async def create_override(
    body: OverrideRequest,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """補打卡（管理員直接建立，自動 approved）"""
    override = Override(
        employee_id=body.employee_id,
        check_type=body.check_type,
        override_at=body.override_at,
        reason=body.reason,
        approved_by=int(admin["sub"]) if admin["sub"].isdigit() else None,
        status="approved",
    )
    db.add(override)

    # 同步寫入一筆 attendance
    att = Attendance(
        employee_id=body.employee_id,
        check_type=body.check_type,
        checked_at=body.override_at,
        is_valid=True,
        note=f"補打卡：{body.reason}",
    )
    db.add(att)
    await db.commit()
    return {"success": True}


@router.get("/override-requests", response_model=list[OverrideRequestOut])
async def list_override_requests(
    status: Optional[str] = None,   # pending | approved | rejected
    employee_id: Optional[int] = None,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """查詢員工補打卡申請列表"""
    q = (
        select(Override, Employee.display_name, Employee.picture_url)
        .join(Employee, Override.employee_id == Employee.id)
        .order_by(Override.created_at.desc())
    )
    if status:
        q = q.where(Override.status == status)
    if employee_id:
        q = q.where(Override.employee_id == employee_id)
    rows = (await db.execute(q)).all()
    return [
        OverrideRequestOut(
            id=ov.id,
            employee_id=ov.employee_id,
            display_name=name,
            picture_url=pic,
            check_type=ov.check_type,
            override_at=ov.override_at.replace(tzinfo=timezone.utc).astimezone(TZ_TAIPEI),
            reason=ov.reason,
            status=ov.status,
            reject_reason=ov.reject_reason,
            created_at=ov.created_at.replace(tzinfo=timezone.utc).astimezone(TZ_TAIPEI),
        )
        for ov, name, pic in rows
    ]


@router.patch("/override-requests/{request_id}/approve")
async def approve_override_request(
    request_id: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """審核通過補打卡申請，並同步寫入出勤紀錄"""
    result = await db.execute(select(Override).where(Override.id == request_id))
    ov = result.scalar_one_or_none()
    if not ov:
        raise HTTPException(status_code=404, detail="申請不存在")
    if ov.status != "pending":
        raise HTTPException(status_code=400, detail="此申請已處理過")

    ov.status = "approved"
    ov.approved_by = int(admin["sub"]) if str(admin["sub"]).isdigit() else None

    # 同步寫入一筆 attendance
    att = Attendance(
        employee_id=ov.employee_id,
        check_type=ov.check_type,
        checked_at=ov.override_at,
        is_valid=True,
        note=f"補打卡：{ov.reason}",
    )
    db.add(att)
    await db.commit()
    return {"success": True}


@router.patch("/override-requests/{request_id}/reject")
async def reject_override_request(
    request_id: int,
    body: OverrideApproveReject,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """駁回補打卡申請"""
    result = await db.execute(select(Override).where(Override.id == request_id))
    ov = result.scalar_one_or_none()
    if not ov:
        raise HTTPException(status_code=404, detail="申請不存在")
    if ov.status != "pending":
        raise HTTPException(status_code=400, detail="此申請已處理過")

    ov.status = "rejected"
    ov.reject_reason = body.reject_reason
    await db.commit()
    return {"success": True}


@router.get("/settings", response_model=SystemSettingsOut)
async def get_settings(
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """取得系統設定"""
    result = await db.execute(select(SystemSettings).where(SystemSettings.id == 1))
    cfg = result.scalar_one_or_none()
    if not cfg:
        cfg = SystemSettings(id=1)
        db.add(cfg)
        await db.commit()
        await db.refresh(cfg)
    return cfg


@router.put("/settings", response_model=SystemSettingsOut)
async def update_settings(
    body: SystemSettingsUpdate,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """更新系統設定"""
    result = await db.execute(select(SystemSettings).where(SystemSettings.id == 1))
    cfg = result.scalar_one_or_none()
    if not cfg:
        cfg = SystemSettings(id=1)
        db.add(cfg)
    if body.gps_enabled is not None:
        cfg.gps_enabled = body.gps_enabled
    if body.office_lat is not None:
        cfg.office_lat = body.office_lat
    if body.office_lng is not None:
        cfg.office_lng = body.office_lng
    if body.office_radius_m is not None:
        cfg.office_radius_m = body.office_radius_m
    await db.commit()
    await db.refresh(cfg)
    return cfg


@router.get("/export/monthly")
async def export_monthly(
    year:  int,
    month: int,
    employee_ids: Optional[str] = None,
    admin: dict = Depends(require_admin),
    db:    AsyncSession = Depends(get_db),
):
    """匯出指定年月的打卡記錄為 Excel；employee_ids 為逗號分隔的員工 id，空白表示全部"""
    # 該月範圍（轉換為 UTC，台灣時間 = UTC+8）
    last_day  = calendar.monthrange(year, month)[1]
    day_start = datetime(year, month, 1) - timedelta(hours=8)
    day_end   = datetime(year, month, last_day, 23, 59, 59) - timedelta(hours=8)

    # 解析員工篩選
    filter_ids: list[int] | None = None
    if employee_ids:
        try:
            filter_ids = [int(x) for x in employee_ids.split(",") if x.strip()]
        except ValueError:
            pass

    # 查詢
    att_query = (
        select(Attendance, Employee.display_name)
        .join(Employee, Attendance.employee_id == Employee.id)
        .where(and_(Attendance.checked_at >= day_start, Attendance.checked_at <= day_end))
    )
    if filter_ids:
        att_query = att_query.where(Attendance.employee_id.in_(filter_ids))
    result = await db.execute(att_query.order_by(Employee.display_name, Attendance.checked_at))
    rows = result.all()

    # 先按員工分成 clock_in / clock_out 兩個清單
    emp_ins:  dict[str, list] = {}
    emp_outs: dict[str, list] = {}
    for att, name in rows:
        local_dt = att.checked_at.replace(tzinfo=timezone.utc).astimezone(TZ_TAIPEI)
        if att.check_type == CheckType.clock_in:
            emp_ins.setdefault(name, []).append(local_dt)
        else:
            emp_outs.setdefault(name, []).append(local_dt)

    # 以上班時間的日期為主，找該次上班後 24 小時內第一筆下班
    data: dict[str, dict] = {}
    for name, ins in emp_ins.items():
        ins.sort()
        outs = sorted(emp_outs.get(name, []))
        used = set()
        data[name] = {}
        for ci in ins:
            d = ci.date()
            data[name].setdefault(d, {"clock_in": None, "clock_out": None})
            if data[name][d]["clock_in"] is None:
                data[name][d]["clock_in"] = ci
            # 找 ci 之後、24 小時內最近的未用下班紀錄
            for i, co in enumerate(outs):
                if i not in used and co > ci and (co - ci).total_seconds() <= 86400:
                    data[name][d]["clock_out"] = co
                    used.add(i)
                    break

    # 員工列表（包含沒打卡的，依篩選條件過濾）
    emp_query = select(Employee).order_by(Employee.display_name)
    if filter_ids:
        emp_query = emp_query.where(Employee.id.in_(filter_ids))
    emp_result = await db.execute(emp_query)
    all_employees = [e.display_name for e in emp_result.scalars().all()]

    # ── 建立 Excel ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月打卡記錄"

    # 顏色
    CLR_HEADER  = "1E293B"
    CLR_DATE    = "334155"
    CLR_PRESENT = "D1FAE5"
    CLR_ABSENT  = "FEE2E2"
    CLR_PARTIAL = "FEF3C7"
    CLR_WEEKEND = "F1F5F9"
    CLR_WHITE   = "FFFFFF"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    WEEKDAY_ZH = ["一", "二", "三", "四", "五", "六", "日"]

    # ── 標題行 ────────────────────────────────────────────────────────────────
    headers = ["員工", "日期", "星期", "上班打卡", "下班打卡", "工時", "狀態"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=CLR_WHITE, size=11)
        cell.fill      = PatternFill("solid", fgColor=CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 22

    # ── 資料行 ────────────────────────────────────────────────────────────────
    row_idx = 2
    for emp_name in all_employees:
        emp_data = data.get(emp_name, {})
        cur = date(year, month, 1)
        end = date(year, month, last_day)
        while cur <= end:
            weekday = cur.weekday()
            is_weekend = weekday >= 5
            rec = emp_data.get(cur)

            ci = rec["clock_in"]  if rec else None
            co = rec["clock_out"] if rec else None

            if ci and co:
                status, bg = "正常", CLR_PRESENT
            elif ci:
                status, bg = "未下班", CLR_PARTIAL
            else:
                status, bg = "休息", CLR_WEEKEND

            # 計算工時
            work_hours = ""
            if ci and co:
                delta = co - ci
                h2 = int(delta.total_seconds() // 3600)
                m2 = int((delta.total_seconds() % 3600) // 60)
                work_hours = f"{h2}:{m2:02d}"

            fill = PatternFill("solid", fgColor=bg)
            values = [
                emp_name,
                cur.strftime("%Y/%m/%d"),
                WEEKDAY_ZH[weekday],
                ci.strftime("%H:%M") if ci else "—",
                co.strftime("%H:%M") if co else "—",
                work_hours or "—",
                status,
            ]
            aligns = ["left", "center", "center", "center", "center", "center", "center"]
            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(horizontal=aln, vertical="center")
                if col == 1:
                    cell.font = Font(bold=True, size=10)
                else:
                    cell.font = Font(size=10)
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            cur += timedelta(days=1)

    # ── 欄寬 ─────────────────────────────────────────────────────────────────
    col_widths = [18, 14, 6, 12, 12, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 凍結首行
    ws.freeze_panes = "A2"

    # ── 輸出 ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/payroll")
async def export_payroll(
    year: int,
    month: int,
    employee_ids: Optional[str] = None,
    columns: Optional[str] = None,   # 逗號分隔，ex: "worked_h,overtime_h,base_pay,overtime_pay,holiday_pay,deductions,total_pay"
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """匯出薪資報表 Excel；支援員工篩選與欄位選擇"""
    filter_ids: list[int] | None = None
    if employee_ids:
        try:
            filter_ids = [int(x) for x in employee_ids.split(",") if x.strip()]
        except ValueError:
            pass

    # 要輸出的欄位（預設全部）
    ALL_COLS = ["worked_h", "overtime_h", "late_h", "early_h", "base_pay", "overtime_pay", "holiday_pay", "deductions", "total_pay", "status"]
    selected = [c.strip() for c in columns.split(",")] if columns else ALL_COLS
    selected = [c for c in selected if c in ALL_COLS]  # 過濾非法欄位

    COL_LABEL = {
        "worked_h":    "出勤(h)",
        "overtime_h":  "加班(h)",
        "late_h":      "遲到(h)",
        "early_h":     "早退(h)",
        "base_pay":    "底薪",
        "overtime_pay":"加班薪",
        "holiday_pay": "特別假日",
        "deductions":  "扣款",
        "total_pay":   "合計",
        "status":      "狀態",
    }

    # 查詢薪資資料
    q = (
        select(PayrollRecord, Employee.display_name)
        .join(Employee, PayrollRecord.employee_id == Employee.id)
        .where(PayrollRecord.year == year, PayrollRecord.month == month)
        .order_by(Employee.display_name)
    )
    if filter_ids:
        q = q.where(PayrollRecord.employee_id.in_(filter_ids))
    result = await db.execute(q)
    rows = result.all()

    # ── Excel ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月薪資"

    CLR_HEADER = "1E293B"
    CLR_WHITE  = "FFFFFF"
    CLR_EVEN   = "F8FAFC"
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["員工"] + [COL_LABEL[c] for c in selected]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=CLR_WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    def col_val(rec, col_key):
        if col_key == "worked_h":    return round(rec.worked_minutes / 60, 1)
        if col_key == "overtime_h":  return round(rec.overtime_minutes / 60, 1)
        if col_key == "late_h":      return round(rec.late_minutes / 60, 1)
        if col_key == "early_h":     return round(rec.early_leave_minutes / 60, 1)
        if col_key == "base_pay":    return rec.base_pay
        if col_key == "overtime_pay":return rec.overtime_pay
        if col_key == "holiday_pay": return rec.holiday_pay
        if col_key == "deductions":  return rec.deductions
        if col_key == "total_pay":   return rec.total_pay
        if col_key == "status":      return "已確認" if rec.status == "finalized" else "未確認"

    for i, (rec, emp_name) in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor=CLR_EVEN if i % 2 == 0 else CLR_WHITE)
        ws.cell(row=i, column=1, value=emp_name).font = Font(bold=True, size=10)
        ws.cell(row=i, column=1).border = border
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center")
        for j, c in enumerate(selected, 2):
            cell = ws.cell(row=i, column=j, value=col_val(rec, c))
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)
        ws.row_dimensions[i].height = 18

    # 欄寬
    ws.column_dimensions["A"].width = 16
    for j in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"payroll_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════════════════
# 管理員帳號管理
# ══════════════════════════════════════════════════════════════════════════

VALID_PERMISSIONS = {"attendance", "employees", "schedule", "export", "salary", "settings"}

@router.get("/admins", response_model=list[AdminUserOut])
async def list_admins(
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """取得所有管理員列表（super_admin 排最前）"""
    result = await db.execute(
        select(AdminUser).order_by(AdminUser.role.desc(), AdminUser.id)
    )
    admins = result.scalars().all()
    return [AdminUserOut.from_orm_with_perms(a) for a in admins]


@router.post("/admins", response_model=AdminUserOut, status_code=201)
async def create_admin(
    body: AdminCreateRequest,
    admin: dict = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    """新增管理員（僅 super_admin 可執行）"""
    existing = await db.execute(
        select(AdminUser).where(AdminUser.username == body.username)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="帳號已存在")

    valid_perms = [p for p in body.permissions if p in VALID_PERMISSIONS]
    new_admin = AdminUser(
        username=body.username,
        hashed_password=pwd_context.hash(body.password),
        display_name=body.display_name,
        role="admin",
        permissions=",".join(valid_perms),
    )
    db.add(new_admin)
    await db.commit()
    await db.refresh(new_admin)
    return AdminUserOut.from_orm_with_perms(new_admin)


@router.patch("/admins/{admin_id}/permissions", response_model=AdminUserOut)
async def update_admin_permissions(
    admin_id: int,
    body: AdminPermissionsUpdate,
    admin: dict = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    """更新管理員權限（僅 super_admin 可執行）"""
    result = await db.execute(select(AdminUser).where(AdminUser.id == admin_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="管理員不存在")
    if target.role.value == "super_admin":
        raise HTTPException(status_code=400, detail="超級管理員權限不可修改")

    valid_perms = [p for p in body.permissions if p in VALID_PERMISSIONS]
    target.permissions = ",".join(valid_perms)
    await db.commit()
    await db.refresh(target)
    return AdminUserOut.from_orm_with_perms(target)


@router.delete("/admins/{admin_id}", status_code=204)
async def delete_admin(
    admin_id: int,
    admin: dict = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    """刪除管理員（僅 super_admin 可執行；super_admin 本身不可被刪除）"""
    result = await db.execute(select(AdminUser).where(AdminUser.id == admin_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="管理員不存在")
    if target.role.value == "super_admin":
        raise HTTPException(status_code=400, detail="超級管理員帳號不可被刪除")

    await db.delete(target)
    await db.commit()


@router.put("/admins/{admin_id}/password")
async def change_admin_password(
    admin_id: int,
    body: AdminPasswordUpdate,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """修改管理員密碼（super_admin 可改任何人；一般管理員只能改自己）"""
    current_id = int(admin["sub"]) if admin["sub"].isdigit() else None
    if admin.get("role") != "super_admin" and current_id != admin_id:
        raise HTTPException(status_code=403, detail="無權修改其他管理員的密碼")

    result = await db.execute(select(AdminUser).where(AdminUser.id == admin_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="管理員不存在")

    target.hashed_password = pwd_context.hash(body.new_password)
    await db.commit()
    return {"message": "密碼已更新"}


@router.delete("/employees/{employee_id}", status_code=204)
async def delete_employee(
    employee_id: int,
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """刪除員工（同時清除其打卡與補打卡記錄）"""
    result = await db.execute(select(Employee).where(Employee.id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="員工不存在")
    await db.execute(delete(Override).where(Override.employee_id == employee_id))
    await db.execute(delete(Attendance).where(Attendance.employee_id == employee_id))
    await db.delete(emp)
    await db.commit()


@router.delete("/attendance/all")
async def clear_all_attendance(
    admin: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """清除所有打卡與補打卡記錄"""
    await db.execute(delete(Override))
    await db.execute(delete(Attendance))
    await db.commit()
    return {"success": True, "message": "所有打卡記錄已清除"}


@router.post("/reset-all", status_code=200)
async def reset_all_data(
    admin: dict = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    清除所有業務資料（僅 super_admin 可執行）。
    保留：admin_users、system_settings。
    刪除：打卡紀錄、薪資設定與紀錄、排班、假別、職位、員工。
    """
    # 依 FK 順序：子表先刪
    await db.execute(delete(PayrollDayOverride))
    await db.execute(delete(LeaveRequest))
    await db.execute(delete(LeaveRecord))
    await db.execute(delete(EmployeeLeaveType))
    await db.execute(delete(PositionLeaveType))
    await db.execute(delete(PayrollRecord))
    await db.execute(delete(EmployeeSalaryConfig))
    await db.execute(delete(ShiftSalaryConfig))
    await db.execute(delete(Schedule))
    await db.execute(delete(Override))
    await db.execute(delete(Attendance))
    # 解除 employees.position_id 循環 FK 再刪
    from sqlalchemy import text
    await db.execute(text("UPDATE employees SET position_id = NULL"))
    await db.execute(delete(Employee))
    await db.execute(delete(Position))
    await db.execute(delete(Shift))
    await db.execute(delete(LeaveType))
    await db.commit()
    return {"success": True, "message": "所有業務資料已清除"}
